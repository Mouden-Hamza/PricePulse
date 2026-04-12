import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os


COLORS = {
    "header_dark": "1a1a2e",
    "header_blue": "16213e",
    "accent": "0f3460",
    "critique": "e74c3c",
    "moyen": "f39c12",
    "faible": "3498db",
    "normal": "2ecc71",
    "light_gray": "f8f9fa",
    "mid_gray": "dee2e6",
    "white": "ffffff",
    "text_dark": "212529",
}

thin = Side(style="thin", color="dee2e6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _cell(ws, row, col, value, bold=False, color=None, bg=None, align="left", num_format=None, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color or COLORS["text_dark"], size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER
    if num_format:
        c.number_format = num_format
    return c


def _header_row(ws, row, cols, bg=None):
    bg = bg or COLORS["header_dark"]
    for col, label in enumerate(cols, 1):
        _cell(ws, row, col, label, bold=True, color=COLORS["white"], bg=bg, align="center", size=10)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(df_analyzed: pd.DataFrame, summary: dict, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_executive(wb, df_analyzed, summary)
    _sheet_anomalies(wb, df_analyzed)
    _sheet_by_category(wb, df_analyzed)
    _sheet_top_opportunities(wb, df_analyzed)
    _sheet_full_catalog(wb, df_analyzed)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Rapport généré : {output_path}")


def _sheet_executive(wb, df, summary):
    ws = wb.create_sheet("Résumé exécutif")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    _cell(ws, 1, 1, "PRICEPULSE — RAPPORT HEBDOMADAIRE", bold=True,
          color=COLORS["white"], bg=COLORS["header_dark"], align="center", size=16)
    ws.merge_cells("A1:H1")

    date_str = datetime.now().strftime("%d/%m/%Y")
    _cell(ws, 2, 1, f"Généré le {date_str} · Catalogue de {summary['total_products']:,} références",
          color="666666", align="center", size=10)
    ws.merge_cells("A2:H2")

    ws.row_dimensions[4].height = 20
    _cell(ws, 4, 1, "INDICATEURS CLÉS", bold=True, color=COLORS["white"],
          bg=COLORS["accent"], align="center", size=11)
    ws.merge_cells("A4:H4")

    kpis = [
        ("Total références", f"{summary['total_products']:,}", COLORS["accent"]),
        ("Anomalies détectées", f"{summary['total_anomalies']:,}", COLORS["critique"]),
        ("Taux d'anomalies", f"{summary['anomaly_rate_pct']}%", COLORS["moyen"]),
        ("Marge moyenne", f"{summary['avg_margin_rate']}%", COLORS["normal"]),
        ("Critiques", f"{summary['critique']}", COLORS["critique"]),
        ("Moyennes", f"{summary['moyen']}", COLORS["moyen"]),
        ("Faibles", f"{summary['faible']}", COLORS["faible"]),
        ("Sur-tarifés", f"{summary['overpriced_count']}", COLORS["moyen"]),
    ]

    for i, (label, value, color) in enumerate(kpis):
        col = i + 1
        _cell(ws, 5, col, label, bold=False, color="555555", bg=COLORS["light_gray"], align="center", size=9)
        _cell(ws, 6, col, value, bold=True, color=color, bg=COLORS["white"], align="center", size=18)

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 36

    _cell(ws, 8, 1, "RÉPARTITION DES ANOMALIES PAR SÉVÉRITÉ", bold=True,
          color=COLORS["white"], bg=COLORS["header_blue"], align="center", size=11)
    ws.merge_cells("A8:D8")

    severity_data = df[df["severity"] != "normal"].groupby("severity").size().reset_index(name="count")
    severity_order = ["critique", "moyen", "faible"]
    severity_data["severity"] = pd.Categorical(severity_data["severity"], categories=severity_order, ordered=True)
    severity_data = severity_data.sort_values("severity")

    headers_sev = ["Sévérité", "Nombre", "% du total anomalies", "Action recommandée"]
    _header_row(ws, 9, headers_sev, bg=COLORS["accent"])

    actions = {
        "critique": "Révision immédiate requise",
        "moyen": "Analyse sous 48h",
        "faible": "Revue hebdomadaire",
    }
    sev_colors = {"critique": COLORS["critique"], "moyen": COLORS["moyen"], "faible": COLORS["faible"]}

    for r, (_, row) in enumerate(severity_data.iterrows(), 10):
        sev = row["severity"]
        pct = round(row["count"] / summary["total_anomalies"] * 100, 1)
        _cell(ws, r, 1, sev.capitalize(), bold=True, color=sev_colors[sev], bg=COLORS["white"])
        _cell(ws, r, 2, row["count"], align="center")
        _cell(ws, r, 3, f"{pct}%", align="center")
        _cell(ws, r, 4, actions[sev])

    _cell(ws, 14, 1, "RÉPARTITION PAR CATÉGORIE", bold=True,
          color=COLORS["white"], bg=COLORS["header_blue"], align="center", size=11)
    ws.merge_cells("A14:D14")

    cat_summary = df.groupby("category").agg(
        total=("product_id", "count"),
        anomalies=("severity", lambda x: (x != "normal").sum()),
        avg_margin=("margin_rate", lambda x: round(x.mean() * 100, 1))
    ).reset_index().sort_values("anomalies", ascending=False)

    _header_row(ws, 15, ["Catégorie", "Références", "Anomalies", "Marge moy. (%)"], bg=COLORS["accent"])
    for r, (_, row) in enumerate(cat_summary.iterrows(), 16):
        _cell(ws, r, 1, row["category"])
        _cell(ws, r, 2, row["total"], align="center")
        _cell(ws, r, 3, row["anomalies"], align="center",
              color=COLORS["critique"] if row["anomalies"] > 50 else COLORS["text_dark"])
        _cell(ws, r, 4, f"{row['avg_margin']}%", align="center")

    _set_col_widths(ws, [22, 16, 16, 16, 20, 16, 16, 22])


def _sheet_anomalies(wb, df):
    ws = wb.create_sheet("Anomalies détectées")
    ws.sheet_view.showGridLines = False

    anomalies = df[df["severity"] != "normal"].sort_values("severity_rank", ascending=False).head(500)

    cols = ["ID Produit", "Nom", "Catégorie", "Prix actuel (€)", "Prix historique (€)",
            "Écart hist. (%)", "Prix concurrent (€)", "Écart conc. (%)",
            "Marge (%)", "Sévérité", "Direction"]
    _header_row(ws, 1, cols, bg=COLORS["header_dark"])

    sev_colors = {"critique": "fde8e8", "moyen": "fef3cd", "faible": "dbeafe"}

    for r, (_, row) in enumerate(anomalies.iterrows(), 2):
        bg = sev_colors.get(row["severity"], COLORS["white"])
        _cell(ws, r, 1, row["product_id"], bg=bg)
        _cell(ws, r, 2, row["product_name"], bg=bg)
        _cell(ws, r, 3, row["category"], bg=bg)
        _cell(ws, r, 4, row["current_price"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 5, row["historical_avg_price"], align="right", num_format="#,##0.00 €", bg=bg)
        ecart_h = row["gap_vs_historical_pct"]
        _cell(ws, r, 6, f"{'+' if ecart_h > 0 else ''}{ecart_h}%", align="center",
              color=COLORS["critique"] if abs(ecart_h) > 20 else COLORS["text_dark"], bg=bg)
        _cell(ws, r, 7, row["competitor_price"], align="right", num_format="#,##0.00 €", bg=bg)
        ecart_c = row["gap_vs_competitor_pct"]
        _cell(ws, r, 8, f"{'+' if ecart_c > 0 else ''}{ecart_c}%", align="center",
              color=COLORS["critique"] if abs(ecart_c) > 20 else COLORS["text_dark"], bg=bg)
        _cell(ws, r, 9, f"{round(row['margin_rate']*100, 1)}%", align="center",
              color=COLORS["critique"] if row["margin_rate"] < 0.05 else COLORS["text_dark"], bg=bg)
        _cell(ws, r, 10, row["severity"].capitalize(), bold=True,
              color={"critique": COLORS["critique"], "moyen": COLORS["moyen"], "faible": COLORS["faible"]}.get(row["severity"]),
              align="center", bg=bg)
        _cell(ws, r, 11, row["price_direction"], align="center", bg=bg)

    _set_col_widths(ws, [12, 20, 16, 14, 16, 13, 16, 13, 10, 12, 16])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(anomalies)+1}"


def _sheet_by_category(wb, df):
    ws = wb.create_sheet("Analyse par catégorie")
    ws.sheet_view.showGridLines = False

    cat_data = df.groupby("category").agg(
        nb_produits=("product_id", "count"),
        anomalies=("severity", lambda x: (x != "normal").sum()),
        critiques=("severity", lambda x: (x == "critique").sum()),
        marge_moy=("margin_rate", lambda x: round(x.mean() * 100, 2)),
        marge_min=("margin_rate", lambda x: round(x.min() * 100, 2)),
        ca_potentiel=("current_price", lambda x: round(x.sum(), 0)),
        marge_totale=("gross_margin_euros", lambda x: round(x.sum(), 0)),
        volume_total=("sales_volume_30d", "sum"),
    ).reset_index().sort_values("anomalies", ascending=False)

    headers = ["Catégorie", "Références", "Anomalies", "dont Critiques",
               "Marge moy. (%)", "Marge min. (%)", "CA catalogue (€)",
               "Marge totale (€)", "Volume 30j"]
    _header_row(ws, 1, headers, bg=COLORS["header_dark"])

    for r, (_, row) in enumerate(cat_data.iterrows(), 2):
        bg = COLORS["light_gray"] if r % 2 == 0 else COLORS["white"]
        _cell(ws, r, 1, row["category"], bold=True, bg=bg)
        _cell(ws, r, 2, row["nb_produits"], align="center", bg=bg)
        _cell(ws, r, 3, row["anomalies"], align="center",
              color=COLORS["critique"] if row["anomalies"] > 50 else COLORS["text_dark"], bg=bg)
        _cell(ws, r, 4, row["critiques"], align="center",
              color=COLORS["critique"] if row["critiques"] > 10 else COLORS["text_dark"], bold=row["critiques"] > 10, bg=bg)
        _cell(ws, r, 5, f"{row['marge_moy']}%", align="center", bg=bg)
        _cell(ws, r, 6, f"{row['marge_min']}%", align="center",
              color=COLORS["critique"] if row["marge_min"] < 0 else COLORS["text_dark"], bg=bg)
        _cell(ws, r, 7, row["ca_potentiel"], align="right", num_format="#,##0 €", bg=bg)
        _cell(ws, r, 8, row["marge_totale"], align="right", num_format="#,##0 €", bg=bg)
        _cell(ws, r, 9, row["volume_total"], align="center", bg=bg)

    _set_col_widths(ws, [20, 12, 12, 14, 13, 13, 18, 18, 12])
    ws.freeze_panes = "A2"


def _sheet_top_opportunities(wb, df):
    ws = wb.create_sheet("Top opportunités")
    ws.sheet_view.showGridLines = False

    _cell(ws, 1, 1, "TOP 50 — PRODUITS SOUS-TARIFÉS (opportunité de hausse)", bold=True,
          color=COLORS["white"], bg=COLORS["normal"], align="center", size=11)
    ws.merge_cells("A1:G1")

    underpriced = df[
        (df["flag_competitor"]) &
        (df["gap_vs_competitor_pct"] < -10) &
        (df["sales_volume_30d"] > 10)
    ].sort_values("gap_vs_competitor_pct").head(50)

    headers_u = ["ID", "Produit", "Catégorie", "Prix actuel (€)", "Prix concurrent (€)",
                 "Écart (%)", "Volume 30j"]
    _header_row(ws, 2, headers_u, bg=COLORS["normal"])

    for r, (_, row) in enumerate(underpriced.iterrows(), 3):
        _cell(ws, r, 1, row["product_id"])
        _cell(ws, r, 2, row["product_name"])
        _cell(ws, r, 3, row["category"])
        _cell(ws, r, 4, row["current_price"], align="right", num_format="#,##0.00 €")
        _cell(ws, r, 5, row["competitor_price"], align="right", num_format="#,##0.00 €")
        _cell(ws, r, 6, f"{row['gap_vs_competitor_pct']}%", align="center", color=COLORS["normal"], bold=True)
        _cell(ws, r, 7, row["sales_volume_30d"], align="center")

    start_over = len(underpriced) + 5
    _cell(ws, start_over, 1, "TOP 50 — PRODUITS SUR-TARIFÉS (risque de perte de compétitivité)", bold=True,
          color=COLORS["white"], bg=COLORS["critique"], align="center", size=11)
    ws.merge_cells(f"A{start_over}:G{start_over}")

    overpriced = df[
        (df["flag_competitor"]) &
        (df["gap_vs_competitor_pct"] > 10) &
        (df["sales_volume_30d"] > 5)
    ].sort_values("gap_vs_competitor_pct", ascending=False).head(50)

    _header_row(ws, start_over + 1, headers_u, bg=COLORS["critique"])

    for r, (_, row) in enumerate(overpriced.iterrows(), start_over + 2):
        _cell(ws, r, 1, row["product_id"])
        _cell(ws, r, 2, row["product_name"])
        _cell(ws, r, 3, row["category"])
        _cell(ws, r, 4, row["current_price"], align="right", num_format="#,##0.00 €")
        _cell(ws, r, 5, row["competitor_price"], align="right", num_format="#,##0.00 €")
        _cell(ws, r, 6, f"{row['gap_vs_competitor_pct']}%", align="center", color=COLORS["critique"], bold=True)
        _cell(ws, r, 7, row["sales_volume_30d"], align="center")

    _set_col_widths(ws, [12, 22, 16, 14, 16, 12, 12])


def _sheet_full_catalog(wb, df):
    ws = wb.create_sheet("Catalogue complet")
    ws.sheet_view.showGridLines = False

    cols = ["ID", "Produit", "Catégorie", "Coût achat (€)", "Prix actuel (€)",
            "Prix historique (€)", "Prix concurrent (€)", "Livraison (€)",
            "Marge (€)", "Marge (%)", "Volume 30j", "Sévérité"]
    _header_row(ws, 1, cols, bg=COLORS["header_dark"])

    sev_colors = {"critique": "fde8e8", "moyen": "fef3cd", "faible": "dbeafe", "normal": COLORS["white"]}

    for r, (_, row) in enumerate(df.iterrows(), 2):
        bg = sev_colors.get(row["severity"], COLORS["white"])
        _cell(ws, r, 1, row["product_id"], bg=bg)
        _cell(ws, r, 2, row["product_name"], bg=bg)
        _cell(ws, r, 3, row["category"], bg=bg)
        _cell(ws, r, 4, row["cost_price"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 5, row["current_price"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 6, row["historical_avg_price"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 7, row["competitor_price"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 8, row["delivery_cost"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 9, row["gross_margin_euros"], align="right", num_format="#,##0.00 €", bg=bg)
        _cell(ws, r, 10, f"{round(row['margin_rate']*100, 1)}%", align="center", bg=bg)
        _cell(ws, r, 11, row["sales_volume_30d"], align="center", bg=bg)
        _cell(ws, r, 12, row["severity"].capitalize(), bold=row["severity"] == "critique",
              color={"critique": COLORS["critique"], "moyen": COLORS["moyen"],
                     "faible": COLORS["faible"], "normal": COLORS["normal"]}.get(row["severity"]),
              align="center", bg=bg)

    _set_col_widths(ws, [12, 20, 16, 14, 14, 16, 16, 12, 12, 10, 11, 12])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(df)+1}"
